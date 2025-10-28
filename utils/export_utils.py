import pandas as pd
import io
import json
from flask import send_file, flash, redirect
from database.reports import get_operating_hours_report, get_status_at_time_report, get_full_history_report
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill
from database.wells_operations import get_all_wells, get_well_maintenance_operations, get_well_by_id
from utils.date_utils import gregorian_to_jalali

def export_operating_hours_to_excel(date_jalali, month_jalali, report_type):
    """خروجی اکسل برای گزارش ساعات کارکرد"""
    results = get_operating_hours_report(date_jalali, month_jalali, report_type)
    
    if not results:
        from flask import flash, redirect
        flash('هیچ داده‌ای برای دانلود وجود ندارد', 'warning')
        return redirect('/report/operating-hours')
    
    df = pd.DataFrame(results)
    
    if report_type == 'daily':
        period_title = f"گزارش ساعات کارکرد روزانه - تاریخ: {date_jalali}"
        filename = f'operating_hours_daily_{date_jalali.replace("/", "-")}.xlsx'
    else:
        period_title = f"گزارش ساعات کارکرد ماهانه - ماه: {month_jalali}"
        filename = f'operating_hours_monthly_{month_jalali.replace("/", "-")}.xlsx'
    
    output = create_excel_with_title(df, period_title, 'گزارش ساعات کارکرد')
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def export_status_report_to_excel(date_jalali, time, display_type):
    """خروجی اکسل برای گزارش وضعیت در زمان خاص"""
    results = get_status_at_time_report(date_jalali, time, display_type)
    
    if not results:
        from flask import flash, redirect
        flash('هیچ داده‌ای برای دانلود وجود ندارد', 'warning')
        return redirect('/report/status-at-time')
    
    df = pd.DataFrame(results)
    period_title = f"گزارش وضعیت پمپ‌ها - تاریخ: {date_jalali} - ساعت: {time}"
    filename = f'status_report_{date_jalali}_{time}.xlsx'.replace('/', '-').replace(':', '-')
    
    output = create_excel_with_title(df, period_title, 'گزارش وضعیت')
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def export_full_history_to_excel(from_date_jalali, to_date_jalali, pump_id):
    """خروجی اکسل برای گزارش تاریخچه کامل"""
    results = get_full_history_report(from_date_jalali, to_date_jalali, pump_id)
    
    if not results:
        from flask import flash, redirect
        flash('هیچ داده‌ای برای دانلود وجود ندارد', 'warning')
        return redirect('/report/full-history')
    
    data = []
    for row in results:
        data.append({
            'شماره پمپ': row['pump_number'],
            'نام پمپ': row['name'],
            'وضعیت': row['action_persian'],
            'تاریخ': row['action_date'],
            'ساعت': row['action_time'],
            'علت': row['reason'],
            'توضیحات': row['notes'] or '',
            'کاربر': row['user_name']
        })
    
    df = pd.DataFrame(data)
    period_title = f"تاریخچه تغییرات - از {from_date_jalali} تا {to_date_jalali}"
    filename = f'full_history_{from_date_jalali}_to_{to_date_jalali}.xlsx'.replace('/', '-')
    
    output = create_excel_with_title(df, period_title, 'تاریخچه تغییرات')
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def create_excel_with_title(df, title, sheet_name):
    """ایجاد فایل اکسل با عنوان"""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        
        # اضافه کردن عنوان
        worksheet.insert_rows(1)
        worksheet['A1'] = title

        # merge title across all dataframe columns (handle variable width)
        if df.shape[1] >= 1:
            last_col = get_column_letter(df.shape[1])
            worksheet.merge_cells(f'A1:{last_col}1')

        # style title
        worksheet['A1'].font = Font(size=14, bold=True)
        worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # auto-size columns based on content
        for idx, col in enumerate(df.columns, start=1):
            col_letter = get_column_letter(idx)
            # compute max length in column (including header)
            try:
                max_length = max(df[col].astype(str).map(len).max(), len(str(col)))
            except Exception:
                max_length = len(str(col))

            adjusted_width = (max_length or 0) + 2
            worksheet.column_dimensions[col_letter].width = adjusted_width

    output.seek(0)

    return output

def create_sample_excel_file():
    """ایجاد فایل نمونه برای آپلود"""
    sample_data = {
        'Pump_Number': [1, 1, 2, 3],
        'Action': ['ON', 'OFF', 'ON', 'OFF'],
        'Date_Jalali': ['1403/07/10', '1403/07/10', '1403/07/11', '1403/07/11'],
        'Time_Jalali': ['08:00', '12:30', '09:15', '17:45'],
        'Reason': ['برنامه ریزی شده', 'قطع برق', 'تعمیرات', 'پایان شیفت'],
        'Notes': ['', 'قطع ۲ ساعته', 'تعویض قطعه', '']
    }
    
    df = pd.DataFrame(sample_data)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Sample', index=False)
    
    output.seek(0)
    
    return send_file(
        output,
        as_attachment=True,
        download_name='pump_history_sample.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def export_wells_to_excel():
    """صادرات اکسل برای مشخصات چاه‌ها (لیست چاه‌ها)"""
    wells = get_all_wells()

    if not wells:
        from flask import flash, redirect
        flash('هیچ چاهی برای دانلود وجود ندارد', 'warning')
        return redirect('/wells')

    data = []
    for w in wells:
        # ensure dict-like access
        row = dict(w)
        created = row.get('created_at')
        created_j = gregorian_to_jalali(created).split(' ')[0] if created else ''

        # Exclude: 'نام چاه', 'شماره پمپ مرتبط', 'وضعیت پمپ', 'تاریخ ایجاد'
        data.append({
            'شماره چاه': row.get('well_number'),
            'موقعیت': row.get('location'),
            'عمق کل': row.get('total_depth'),
            'عمق نصب پمپ': row.get('pump_installation_depth'),
            'قطر چاه': row.get('well_diameter'),
            'برند پمپ': row.get('current_pump_brand'),
            'مدل پمپ': row.get('current_pump_model'),
            'قدرت پمپ': row.get('current_pump_power'),
            'جنس لوله': row.get('current_pipe_material'),
            'قطر لوله': row.get('current_pipe_diameter'),
            'متراژ لوله (متر)': row.get('current_pipe_length_m'),
            'کابل اصلی': row.get('main_cable_specs'),
            'کابل چاه': row.get('well_cable_specs'),
            'مشخصات تابلو': row.get('current_panel_specs'),
            'وضعیت': row.get('status')
        })

    df = pd.DataFrame(data)
    period_title = "مشخصات چاه‌ها"
    filename = f'wells_list_{pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    output = create_excel_with_title(df, period_title, 'چاه‌ها')
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def export_well_history_to_excel(well_id):
    """Export maintenance/history for a single well to Excel.
    Columns: History ID, Well Number, Well Name, Operation Type, Operation Date (Jalali), Performed By,
    [well fields...], Notes. Cells corresponding to changed fields are highlighted.
    """
    operations = get_well_maintenance_operations(well_id, limit=10000)

    if not operations:
        flash('هیچ سابقه‌ای برای این چاه موجود نیست', 'warning')
        return redirect(f'/wells/{well_id}/maintenance')

    # Define the well fields to include (in this order)
    well_fields = [
        ('total_depth', 'عمق کل'),
        ('pump_installation_depth', 'عمق نصب پمپ'),
        ('well_diameter', 'قطر چاه'),
        ('current_pump_brand', 'برند پمپ'),
        ('current_pump_model', 'مدل پمپ'),
        ('current_pump_power', 'قدرت پمپ'),
        ('current_pipe_material', 'جنس لوله'),
        ('current_pipe_diameter', 'قطر لوله'),
        ('current_pipe_length_m', 'متراژ لوله (متر)'),
        ('main_cable_specs', 'کابل اصلی'),
        ('well_cable_specs', 'کابل چاه'),
        ('current_panel_specs', 'مشخصات تابلو'),
        ('status', 'وضعیت')
    ]

    data = []
    changed_lists = []  # parallel list of changed_fields per row

    # try to get well basic info
    well = get_well_by_id(well_id, include_pump_info=False)
    well_number = well['well_number'] if well else ''
    well_name = well['name'] if well else ''

    for op in operations:
        op_dict = dict(op)
        # parse full_snapshot if present
        snapshot = {}
        try:
            if op_dict.get('full_snapshot'):
                snapshot = json.loads(op_dict.get('full_snapshot'))
        except Exception:
            snapshot = {}

        # operation date jalali
        op_date = op_dict.get('operation_date')
        if op_date:
            try:
                jal = gregorian_to_jalali(str(op_date))
                jal_date = jal.split(' ')[0]
            except Exception:
                jal_date = op_date
        else:
            jal_date = ''

        changed = []
        try:
            changed = json.loads(op_dict.get('changed_fields') or '[]')
        except Exception:
            changed = []

        changed_lists.append(changed)

        row = {
            'History ID': op_dict.get('id'),
            'شماره چاه': well_number,
            'نوع عملیات': op_dict.get('operation_type'),
            'تاریخ عملیات': jal_date,
            'انجام دهنده': op_dict.get('performed_by') or ''
        }

        # include well fields from snapshot
        for key, label in well_fields:
            val = snapshot.get(key) if isinstance(snapshot, dict) else ''
            row[label] = val if val is not None else ''

        # notes / reason
        row['یادداشت'] = op_dict.get('reason') or ''

        data.append(row)

    df = pd.DataFrame(data)

    # filename
    filename = f'well_{well_number}_history_{pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

    output = io.BytesIO()
    highlight_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sheet_name = 'تاریخچه تعمیرات'
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # Insert title row
        worksheet.insert_rows(1)
        worksheet['A1'] = f"تاریخچه تعمیرات چاه {well_number}"
        last_col = get_column_letter(df.shape[1]) if df.shape[1] >= 1 else 'A'
        worksheet.merge_cells(f'A1:{last_col}1')
        worksheet['A1'].font = Font(size=14, bold=True)
        worksheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

        # auto-size
        for idx, col in enumerate(df.columns, start=1):
            col_letter = get_column_letter(idx)
            try:
                max_length = max(df[col].astype(str).map(len).max(), len(str(col)))
            except Exception:
                max_length = len(str(col))
            worksheet.column_dimensions[col_letter].width = (max_length or 0) + 2

        # Apply highlights for changed fields
        # header row is now at row 2, data starts at row 3
        header_to_col = {str(df.columns[i]): i + 1 for i in range(len(df.columns))}

        for row_idx, changed in enumerate(changed_lists, start=0):
            excel_row = row_idx + 3
            for field_key in changed:
                # map field_key to Persian header if present
                # find matching header label from well_fields
                label = None
                for k, lab in well_fields:
                    if k == field_key:
                        label = lab
                        break
                # also handle notes and other direct fields
                if field_key == 'notes' or field_key == 'reason':
                    label = 'یادداشت'

                if label and label in header_to_col:
                    col_idx = header_to_col[label]
                    try:
                        cell = worksheet.cell(row=excel_row, column=col_idx)
                        cell.fill = highlight_fill
                    except Exception:
                        pass

    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )